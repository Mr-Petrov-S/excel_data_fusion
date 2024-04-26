import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from regions import list_of_regions
from cancer_loc import men_cancer, women_cancer
import re

# Функция для заполнения федерального округа
def district(a):
    # убираем лишние символы
    a = a.strip().replace('\n', '').replace('\r', '')
    # перебираем и ищем в словаре совпадения
    for dictionary in list_of_regions:
        for key, value in dictionary.items():
            if a in value:
                return key           

# Создаем два списка, соп и зно. Также исключаем таблицы которые не нужны по условию
def two_lists(list_of_files):
    list_of_sop = []
    list_of_zno = []
    # ищем по ключевым словам "сост(состояние)" и "зло(злокачественные)". Выбираем из названия номер таблицы и убираем не нужные по условию
    for file in list_of_files:
        if ('табл' in file.lower()) and ('сост' in file.lower()) and (int(re.findall(r'_\d{3}_', file)[0].strip('_')) > 23) and (int(re.findall(r'_\d{3}_', file)[0].strip('_')) != 57):
            list_of_sop.append(file.split('/')[-1])
        elif ('табл' in file.lower()) and ('зло' in file.lower()) and (int(re.findall(r'_\d{3}_', file)[0].strip('_')) > 11) and (int(re.findall(r'_\d{3}_', file)[0].strip('_')) != 65):
            list_of_zno.append(file.split('/')[-1])

    return list_of_sop, list_of_zno

# Обработка таблиц СОП
def preprocessor_sop(file_list):
    # создаем датафреймы для хранения итогового результата
    table1 = pd.DataFrame()
    table2 = pd.DataFrame()
    table3 = pd.DataFrame()
    # переменные для фильтрации таблиц по типу содержимого внутри
    a = 'Сведения о контингенте больных со злокачественными новообразованиями, состоящем на учете в онкологических учреждениях в 2021 г.'
    b = 'Сведения о лечении злокачественных новообразований (зно), впервые зарегистрированных в 2021 г., подлежащих радикальному лечению'
    c = 'Показатели диагностики злокачественных новообразований, выявленных в 2021 г.'
    # перебираем все таблицы из списка
    for item in file_list:
        # печатаем для простоты отладки в случае ошибок (можно сразу понять на какой таблице что-то пошло не так)
        print(item)
        # загружаем Excel файл
        wb = load_workbook(item)
        # итерация по листам внутри книги excel
        for idx, sheet in enumerate(wb):
            # печатем индекс для понимания на каком мы сейчас листе (также для отладки)
            print(idx)
            # Берем для формирования данных в столбцах
            cell_value = sheet['A1'].value
            # убираем лишние символы
            cell_value = cell_value.replace('\n', ' ').strip()
            cell_value = re.sub(" +", " ", cell_value)
            # с помощью регулярных выражений получаем нужныеданные из строк заголовка таблицы
            # вырезаем найденное для того чтобы в итоговом значении cell_value осталось локализация болезни
            try:
                year = re.findall(r'(20\d{2})', cell_value)[0]
            except IndexError:
                year = 'unknown'
        
            try:
                ind = re.findall(r'(?s)\A.*\ЛЕЧЕНИЮ', cell_value)[0].capitalize()
                cell_value = re.sub(r'(?s)\A.*\ЛЕЧЕНИЮ', '', cell_value).strip()
            except IndexError:
                try:
                    ind = re.findall(r'(?s)\A.*\Г\.', cell_value)[0].capitalize()
                    cell_value = re.sub(r'(?s)\A.*\Г\.', '', cell_value).strip()
                except IndexError:
                    ind = 'unknown'
            # здесь много разных вариантов попадается в таблицах, обрабатываем их все. Порядок важен!
            try:
                table = re.findall(r'(?s)\Продолжение\sтаблиц.\s.*\d{2,4}', cell_value, re.IGNORECASE)[0]
                cell_value = re.sub(r'(?s)\Продолжение\sтаблиц.\s.*\d{2,4}', '', cell_value, re.IGNORECASE).strip()
            except IndexError:
                try:
                    table = re.findall(r'(?s)\Таблица\s\d{2,4}.*\Продолжение', cell_value, re.IGNORECASE)[0]
                    cell_value = cell_value.lower()
                    cell_value = re.sub(r'(?s)\таблица\s\d{2,4}.*\продолжение', '', cell_value, re.IGNORECASE).strip()
                except IndexError:
                    try:
                        table = re.findall(r'(?s)\Таблица.*\d{2}\s', cell_value)[0]
                        cell_value = re.sub(r'(?s)\Таблица.*\d{2}\s', '', cell_value).strip()
                    except IndexError:
                        try:
                            table = re.findall(r'(?s)\Таблица.*\d{2}\.', cell_value)[0]
                            cell_value = re.sub(r'(?s)\Таблица.*\d{2}\.', '', cell_value).strip()
                        except IndexError:
                               try:
                                   table = re.findall(r'(?s)\Таблица.*\d{2,4}', cell_value)[0]
                                   cell_value = re.sub(r'(?s)\Таблица.*\d{2,4}', '', cell_value).strip()
                               except IndexError:
                                    table = 'unknown'          

            # собираем данные для имен колонок
            column_names_sop = []
            # выбираем имена колонок в зависимости от наполнения таблицы
            # флаг для понимания какой тип таблицы сейчас обрабатывается, чтобы правильно сделать олбъединение после обработки
            if ind == a:
                flag = 1
                column_names_sop = [
                    'region',
                    'Взято на учет больных с впервые в жизни уст. диагнозом ЗНО',
                    'в т.ч. выявлены активно, %',
                    'Находились на учете на конец года абсолютное число',
                    'Находились на учете на конец года на 100 тыс. населения',
                    'из них 5 лет и более абсолютное число',
                    'из них 5 лет и более % от сост. на учете',
                    'Индекс накопления контингентов',
                    'Летальность, %'
                ]
            elif ind == b:
                flag = 2
                column_names_sop = [
                    'region',
                    'Число ЗНО, выявленных в отчетном году, радикальное лечение которых закончено в отчетном году',
                    'Число ЗНО, выявленных в отчетном году, радикальное лечение которых % от впервые выявленных',
                    'Число ЗНО, выявленных в отчетном году, радикальное лечение которых будет продолжено (не закончено)',
                    'Число ЗНО, выявленных в отчетном году, радикальное лечение которых % от впервые выявленных',
                    'В том числе с использованием методов только хирургического, %',
                    'В том числе с использованием методов только лучевого, %',
                    'В том числе с использованием методов только лекарственного, %',
                    'В том числе с использованием методов комбинир. или компл. (кроме химиолучевого), %',
                    'В том числе с использованием методов химиолучевого, %'
                ]
            elif ind == c:
                flag = 3
                column_names_sop = [
                    'region',
                    'Зарегистрировано ЗНО (без учтенных посмертно)',
                    'из них диагноз подтвержден морфологически, %',
                    'из них имели стадию заболевания, % I',
                    'из них имели стадию заболевания, % II',
                    'из них имели стадию заболевания, % III',
                    'из них имели стадию заболевания, % IV',
                    'из них имели стадию заболевания, % не установлена',
                    'Летальность на первом году с момента уст. диагноза, %'
                ]

            # Создаем пустой список для данных
            data = []

            
            for row in sheet.iter_rows(min_row=6, values_only=True):
            # Фильтрация пустых ячеек в каждой строке
                filtered_row = [cell for cell in row if cell is not None]
    
                # Добавляем отфильтрованную строку в список, если она не пуста
                if filtered_row:
                    data.append(filtered_row)

            # Закрываем Excel-файл
            wb.close()

            # Преобразовываем список данных в DataFrame
            df = pd.DataFrame(data, columns=column_names_sop)
            # Удаляем все строки с фед.округами из таблицы
            df = df[~df['region'].str.contains('ФО', na=False)]
            df = df.reset_index(drop=True)
        
        
            # добавляем колонку с фед.округами второй по счету с помощью функции, предварительно проверяем не попали ли случайно цифры в строку с регионом
            df = df[~df['region'].apply(lambda x: isinstance(x, int))]
            df.insert(loc=1, column='federal', value=df['region'].apply(lambda x: district(x))) # ламбду можно удалить
            # присваиваем в колонки полученные с помощью регулярок значения
            df['ind'] = ind
            df['year'] = year
            df['loc'] = cell_value
            df['table'] = table
            df['bzz'] = 'СОП'
            # объединяем датафреймы
            # на основе флага объединяем с нужным датафреймом 
            if flag == 1:
                table1 = pd.concat([table1, df], ignore_index=True)
            elif flag == 2:
                table2 = pd.concat([table2, df], ignore_index=True)
            elif flag == 3:
                table3 = pd.concat([table3, df], ignore_index=True)
            
    # возвращаем итоговые датафреймы после всех итераций    
    return table1, table2, table3


# подфункция для функции обработки таблиц зно
# на входе выбирается лист, стартовая строка обработки, максимальная строка обработки(если требуется), минимальная и максимальная колонки.
def create_df(sheet, min_row, max_row=None, min_col=2, max_col=None):
        # листы для хранения значений, один лист для регионов (так как это всегда первая колонка)
        # второй лист для данных которые повторяются 3 раза по 4 колонки
        region_list = []
        data_list = []
        # цикл прохода по строкам для сбора регионов и добавляения их в список
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, max_col=1, values_only=True):
            # Фильтрация пустых ячеек в каждой строке
            filtered_row = [cell for cell in row if cell is not None]
            if filtered_row:
                 region_list.append(filtered_row) 
        # цикл для сбора остальных данных   
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            # Фильтрация пустых ячеек в каждой строке
            filtered_row = [cell for cell in row if cell is not None]
            if filtered_row:
                data_list.append(filtered_row)
        # имена колонок
        column_regions = ['region']
        column_values = ['abs', 'rude_100', 'standart_100', 'error_100']
        # преобразуем в датафреймы
        region_df= pd.DataFrame(region_list, columns=column_regions)
        data_df = pd.DataFrame(data_list, columns=column_values)
        # объединяем по горизонтали
        df = pd.concat([region_df, data_df], axis=1)
        # возвращаем объединенный датафрейм
        return df

# подфункция для функции обработки таблиц зно
# получаем данные из заголовков таблиц
def title_name(sheet):
    title_list = []
    # итерация по строкам до 4-й в которых рапологаются нужные нам данные
    for row in sheet.iter_rows(max_row=4, values_only=True):
        # Фильтрация пустых ячеек в каждой строке
        filtered_row = [cell for cell in row if cell is not None]
        if filtered_row:
            title_list.append(filtered_row)

    for item in title_list:
        # объединяем подстроки в строку
        concatenated_item = ' '.join(map(str, item))
        # Приведение к нижнему регистру 
        concatenated_item_lower = concatenated_item.lower()  
        # Разделение строки
        item_parts = concatenated_item.split()  
        # проверяем вхождение ключевых слов и присваеваем нужные нам значения в переменные
        if 'таблица' in concatenated_item_lower:
            table = concatenated_item

        if 'заболеваемость' in concatenated_item_lower or 'смертность' in concatenated_item_lower:
                ind = item_parts[0].capitalize()

        if 'год' in concatenated_item_lower:
                year = item_parts[1]

        if 'локализация' in concatenated_item_lower:
                loc = ' '.join(item_parts[1:])
    # возвращаем значения переменных    
    return ind, year, loc, table

# функция обработки таблиц ЗНО
def preprocessor_zno(file_list):
    # датафреймы для конечного результата
    result, data_both, data_m, data_f = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    # итерируемся по файлам в листе
    for item in file_list:
        # печатем имена таблиц которые сейчас в работе (для упрощения отладки)
        print(item)
        wb = load_workbook(item)
        # проходим по листам внутри книги
        for sheet in wb.worksheets:
            # выводим имя листа (для отладки)
            print(sheet.title)       
            # применяем функции к разным диапазонам колонок для получения разных данных в зависимости от расположения    
            df_both = create_df(sheet=sheet, min_row=10, min_col=2, max_col=5)
            df_m = create_df(sheet=sheet, min_row=10, min_col=6, max_col=9)
            df_f = create_df(sheet=sheet, min_row=10, min_col=10, max_col=13)
            # аналогично с переменными (нужно чтобы далее присвоить значения в колонки)
            df_both[['ind', 'year', 'loc', 'table']] = title_name(sheet)
            df_m[['ind', 'year', 'loc', 'table']] = title_name(sheet)
            df_f[['ind', 'year', 'loc', 'table']] = title_name(sheet)


        # Закрываем Excel-файл
        wb.close()
        # так как среди наших данных есть чисто мужские и женские болезни, проверяем этот момент
        # в таблицах с мужскими и женскими болезнями меньше столбцов, по логике работы нашего обработчика такие значения могут попасть только в df_both
        # так как этот фрейм собирает в себя первые 5 столбцов. 
        # таким образом если в нем присутствуют значения из ране созданных списков men_cancer и women_cancer, нам нужно присвоить значения в соответствующий фрейм данных
        # и удалить остальные
        if df_both['loc'].isin(men_cancer).any():
            df_m = df_both[df_both['loc'].isin(men_cancer)]
            df_both = pd.DataFrame()
            df_f = pd.DataFrame()
            # так как два других датафрейма будут пустыми, обработке подвергается только один
            # убираем строки с фед округами
            df_m = df_m[~df_m['region'].str.contains('ФО', na=False)].reset_index(drop=True)
            # добавляем фед округ в отдельный столбец
            df_m.insert(loc=1, column='federal', value=df_m['region'].apply(district))
            # добавляем пол в отдельный столбец
            df_m.insert(loc=2, column='gender', value='М')
            # добавялем тип
            df_m['bzz'] = 'ЗНО'
        # аналогично для женских болезней
        elif df_both['loc'].isin(women_cancer).any():
            df_f = df_both[df_both['loc'].isin(women_cancer)]
            df_both = pd.DataFrame()
            df_m = pd.DataFrame()

            df_f = df_f[~df_f['region'].str.contains('ФО', na=False)].reset_index(drop=True)
            df_f.insert(loc=1, column='federal', value=df_f['region'].apply(district))
            df_f.insert(loc=2, column='gender', value='Ж')
            df_f['bzz'] = 'ЗНО'
        # аналогичная обработка для общих болезней
        else:
            df_both = df_both[~df_both['region'].str.contains('ФО', na=False)].reset_index(drop=True)
            df_m = df_m[~df_m['region'].str.contains('ФО', na=False)].reset_index(drop=True)
            df_f = df_f[~df_f['region'].str.contains('ФО', na=False)].reset_index(drop=True)
            
            df_both.insert(loc=1, column='federal', value=df_both['region'].apply(district))
            df_m.insert(loc=1, column='federal', value=df_m['region'].apply(district))
            df_f.insert(loc=1, column='federal', value=df_f['region'].apply(district))

            df_both.insert(loc=2, column='gender', value='Оба пола')
            df_m.insert(loc=2, column='gender', value='М')
            df_f.insert(loc=2, column='gender', value='Ж')

            df_both['bzz'], df_m['bzz'], df_f['bzz'] = ['ЗНО'] * 3
        # добавляем обработанные на этой итерации датафреймы к итоговым
        data_both = pd.concat([data_both, df_both], ignore_index=True)
        data_m = pd.concat([data_m, df_m], ignore_index=True)
        data_f = pd.concat([data_f, df_f], ignore_index=True)
    # последовательно соединяем полученные итоговые датафреймы
    result = pd.concat([data_both, data_m, data_f], ignore_index=True)
    #  возвращаем датафрейм
    return result          

# функция для сохранения файлов
def save_files(dataframe, file_name):
    # путь сохранения
    load_path = 'C:/processed_files/'
    # создаем если его нет
    if not os.path.exists(load_path):
        os.makedirs(load_path, exist_ok=True)

    # Полный путь к файлу результату
    result_file = os.path.join(load_path, f"{file_name}.xlsx")
    # выгружаем файл в excel
    dataframe.to_excel(result_file, index=False)
